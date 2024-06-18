import openpyxl

# Load the source workbook and select the source sheet
source_workbook = openpyxl.load_workbook("Book2.xlsx")
source_sheet = source_workbook.active  # Use .active to get the active sheet or specify sheet name

# Create a new workbook for the destination file
destination_workbook = openpyxl.Workbook()
destination_sheet = destination_workbook.active


row_number = 1

# Extract values from the row
row_values = []
for cell in source_sheet[row_number]:
    row_values.append(cell.column)
print(row_values)

"""
# Iterate over the source sheet and copy each cell's value to the destination sheet
for row in source_sheet.iter_rows(min_row=2, max_row=5, min_col=1, max_col=1):
    for cell in row:
        print(cell.value)
"""
# from openpyxl import load_workbook

# # Load the Excel workbook and select a sheet
# workbook = load_workbook('Book2.xlsx')
# sheet = workbook.active  # You can also specify a sheet name here

# # Get the column names (assuming the first row contains headers)
# column_names = [cell.value for cell in sheet[1]]

# print(column_names)
