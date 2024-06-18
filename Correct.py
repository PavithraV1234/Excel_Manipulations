import openpyxl as opl

while(True):
    print("Menu:")
    print("1.Continue adding files")
    print("2.Exit")
    choice=int(input("Enter your choice:"))
    if choice==1:
        fname=str(input("Enter file name(without extension):"))
        no_of_cols=int(input("Enter the no. of columns to be selected:"))
        col_list=[]
        for i in range(no_of_cols):
            col_name=str(input("Enter column name:"))
            col_list.append(col_name)
        source=opl.load_workbook(fname+".xlsx")
        source_sheet = source.active 
        destn = opl.Workbook("Book3.xlsx")
        destn_sheet=destn.active
    
        row_no=1
        row_values = []
        for i in col_list:
            for cell in source_sheet[row_no]:
                if i==cell.value.strip():
                    row_values.append(cell.column)
        for k in row_values:
            for row in source_sheet.iter_rows(min_row=2, max_row=5, min_col=k, max_col=k):
                for cell in row:
                    print(cell.value)
                 
            

